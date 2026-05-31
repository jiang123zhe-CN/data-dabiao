import io
import json
from typing import BinaryIO

import openpyxl
from sqlalchemy.orm import Session

from app.models.field import Field, ImportRecord
from app.models.user import User

TEMPLATE_HEADERS = [
    "field_code", "name", "english_name", "data_type", "length",
    "precision", "table_name", "database_name", "business_domain",
    "sensitivity_level(自动)", "description", "business_rules",
]


def generate_template() -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "字段导入模板"
    ws.append(TEMPLATE_HEADERS)
    ws.append(["F001", "客户名称", "customer_name", "VARCHAR", 100, None, "dim_customer", "ods", "客户域", "", "客户姓名", ""])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def parse_excel(file: BinaryIO) -> list[dict]:
    wb = openpyxl.load_workbook(file, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    headers = [h.value for h in ws[1]]

    results = []
    for row_idx, row in enumerate(rows, start=2):
        if all(v is None for v in row):
            continue
        row_dict = {}
        for i, header in enumerate(["field_code", "name", "english_name", "data_type", "length",
                                     "precision", "table_name", "database_name", "business_domain",
                                     "sensitivity_level", "description", "business_rules"]):
            val = row[i] if i < len(row) else None
            row_dict[header] = str(val).strip() if val is not None else None
        row_dict["_row"] = row_idx
        results.append(row_dict)

    wb.close()
    return results


def import_fields(db: Session, file: BinaryIO, user: User, file_name: str, file_size: int) -> ImportRecord:
    rows = parse_excel(file)
    total = len(rows)
    errors = []
    success = 0
    new_field_ids = []

    import_record = ImportRecord(
        user_id=user.id,
        file_name=file_name,
        file_size=file_size,
        total_rows=total,
        success_rows=0,
        failed_rows=0,
        status="processing",
    )
    db.add(import_record)
    db.flush()

    for row in rows:
        rownum = row.pop("_row")
        row_errors = []

        if not row.get("field_code"):
            row_errors.append({"row": rownum, "field": "field_code", "message": "字段编码不能为空"})
        if not row.get("name"):
            row_errors.append({"row": rownum, "field": "name", "message": "字段名称不能为空"})
        if not row.get("data_type"):
            row_errors.append({"row": rownum, "field": "data_type", "message": "数据类型不能为空"})
        if not row.get("table_name"):
            row_errors.append({"row": rownum, "field": "table_name", "message": "表名不能为空"})

        # sensitivity_level is optional - system auto-classifies
        sl = row.get("sensitivity_level") or "L2"

        existing = db.query(Field).filter(Field.field_code == row["field_code"]).first()
        if existing:
            row_errors.append({"row": rownum, "field": "field_code", "message": "字段编码已存在"})

        length_val = None
        if row.get("length"):
            try:
                length_val = int(row["length"])
            except ValueError:
                row_errors.append({"row": rownum, "field": "length", "message": "长度必须为整数"})

        precision_val = None
        if row.get("precision"):
            try:
                precision_val = int(row["precision"])
            except ValueError:
                row_errors.append({"row": rownum, "field": "precision", "message": "精度必须为整数"})

        if row_errors:
            errors.extend(row_errors)
            continue

        field = Field(
            field_code=row["field_code"],
            name=row["name"],
            english_name=row.get("english_name"),
            data_type=row["data_type"],
            length=length_val,
            precision=precision_val,
            table_name=row["table_name"],
            database_name=row.get("database_name"),
            business_domain=row.get("business_domain"),
            sensitivity_level=sl,
            description=row.get("description"),
            business_rules=row.get("business_rules"),
            source="excel_import",
            import_batch_id=import_record.id,
            created_by=user.id,
        )
        db.add(field)
        db.flush()
        new_field_ids.append(field.id)
        success += 1

    import_record.success_rows = success
    import_record.failed_rows = total - success
    import_record.status = "completed" if success == total else ("partial" if success > 0 else "failed")
    import_record.error_details = json.dumps(errors, ensure_ascii=False) if errors else None
    db.commit()
    db.refresh(import_record)

    # Auto-classify all newly imported fields via rule engine
    if new_field_ids:
        from app.services.rule_engine import RuleEngine
        engine = RuleEngine(db)
        for fid in new_field_ids:
            field = db.query(Field).filter(Field.id == fid).first()
            if field:
                result = engine.classify_field(field)
                if result["category_id"] or result["tier_level"]:
                    field.classification_id = result.get("category_id") or field.classification_id
                    field.sensitivity_level = result.get("tier_level") or field.sensitivity_level
                    field.tagging_method = "rule_engine"
                    field.tagging_confidence = result.get("confidence", 0.0)
        db.commit()

    return import_record


def export_mappings_to_excel(mappings: list[dict]) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "资产映射数据"

    headers = [
        "映射ID", "目录路径", "目录编码", "字段编码", "字段名称",
        "数据类型", "来源表", "映射来源", "置信度", "创建时间",
    ]
    ws.append(headers)

    for m in mappings:
        ws.append([
            m.get("id"),
            m.get("directory_path", ""),
            m.get("directory_code", ""),
            m.get("field_code", ""),
            m.get("field_name", ""),
            m.get("field_data_type", ""),
            m.get("field_table", ""),
            "AI建议" if m.get("mapping_source") == "ai_suggested" else "手动映射",
            m.get("confidence"),
            m.get("created_at"),
        ])

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_fields_to_excel(fields: list[Field]) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "数据字段"
    ws.append(TEMPLATE_HEADERS + ["status", "is_anomaly"])

    for f in fields:
        ws.append([
            f.field_code, f.name, f.english_name, f.data_type, f.length,
            f.precision, f.table_name, f.database_name, f.business_domain,
            f.sensitivity_level, f.description, f.business_rules,
            f.status, str(f.is_anomaly),
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
