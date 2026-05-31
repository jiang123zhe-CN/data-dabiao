from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.core.database import get_db
from app.core.security import get_current_user, require_role
from app.models.user import User
from app.models.field import Field
from app.models.directory import Directory
from app.models.mapping import DirectoryFieldMapping
from app.models.review_record import ReviewRecord
from app.models.standard import ClassificationCategory, TieringRule
from app.models.tagging import TaggingHistory
from app.models.operation_log import OperationLog
from app.services.excel_service import export_fields_to_excel

router = APIRouter(prefix="/api/reports", tags=["Reports"])


@router.get("/summary")
def get_summary(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    total_fields = db.query(Field).filter(Field.status == "active").count()
    total_dirs = db.query(Directory).filter(Directory.is_active == True).count()
    total_mappings = db.query(DirectoryFieldMapping).count()
    pending_reviews = db.query(ReviewRecord).filter(ReviewRecord.review_status == "pending").count()
    anomaly_count = db.query(Field).filter(Field.is_anomaly == True).count()
    classified_count = db.query(Field).filter(Field.status == "active", Field.classification_id.isnot(None)).count()

    return {
        "total_fields": total_fields,
        "total_directories": total_dirs,
        "total_mappings": total_mappings,
        "pending_reviews": pending_reviews,
        "anomaly_count": anomaly_count,
        "classified_count": classified_count,
        "coverage_pct": round(classified_count / total_fields * 100, 1) if total_fields else 0,
    }


@router.get("/by-directory")
def get_by_directory(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    dirs = db.query(Directory).filter(Directory.is_active == True).all()
    result = []
    for d in dirs:
        count = db.query(DirectoryFieldMapping).filter(DirectoryFieldMapping.directory_id == d.id).count()
        result.append({"directory_id": d.id, "name": d.name, "code": d.code, "field_count": count})
    return result


@router.get("/by-sensitivity")
def get_by_sensitivity(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    from sqlalchemy import func
    result = db.query(Field.sensitivity_level, func.count(Field.id)).filter(Field.status == "active").group_by(Field.sensitivity_level).all()
    return [{"sensitivity_level": r[0], "count": r[1]} for r in result]


@router.get("/export/fields")
def export_fields_report(
    db: Session = Depends(get_db),
    _: User = Depends(require_role("admin")),
):
    fields = db.query(Field).all()
    output = export_fields_to_excel(fields)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=data_fields_report.xlsx"},
    )


@router.get("/export/mappings")
def export_mappings_report(
    db: Session = Depends(get_db),
    _: User = Depends(require_role("admin")),
):
    import io
    import openpyxl

    mappings = db.query(DirectoryFieldMapping).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "映射关系"
    ws.append(["映射ID", "字段ID", "字段编码", "字段名称", "目录ID", "目录名称", "映射来源", "置信度", "创建时间"])

    for m in mappings:
        field = db.query(Field).filter(Field.id == m.field_id).first()
        directory = db.query(Directory).filter(Directory.id == m.directory_id).first()
        ws.append([
            m.id, m.field_id,
            field.field_code if field else "", field.name if field else "",
            m.directory_id, directory.name if directory else "",
            m.mapping_source, m.confidence,
            m.created_at.isoformat() if m.created_at else "",
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=mappings_report.xlsx"},
    )


# ── Compliance Report Endpoints ──

@router.get("/compliance/summary")
def compliance_summary(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    total = db.query(Field).filter(Field.status == "active").count()
    classified = db.query(Field).filter(Field.status == "active", Field.classification_id.isnot(None)).count()
    tiered = db.query(Field).filter(Field.status == "active", Field.sensitivity_level.isnot(None)).count()
    auto_tagged = db.query(Field).filter(Field.status == "active", Field.tagging_method.isnot(None), Field.tagging_method != "manual").count()
    manual_tagged = db.query(Field).filter(Field.status == "active", Field.tagging_method == "manual").count()
    today_ops = db.query(OperationLog).filter(
        OperationLog.created_at >= func.date('now')
    ).count()

    return {
        "total_fields": total,
        "classified_count": classified,
        "tiered_count": tiered,
        "coverage_pct": round(classified / total * 100, 1) if total else 0,
        "auto_tagged": auto_tagged,
        "manual_tagged": manual_tagged,
        "today_operations": today_ops,
    }


@router.get("/compliance/by-category-tier")
def compliance_by_category_tier(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    fields = db.query(Field).filter(Field.status == "active").all()
    categories = {c.id: c.name for c in db.query(ClassificationCategory).filter(ClassificationCategory.is_active == True).all()}

    matrix = {}
    for f in fields:
        cat_name = categories.get(f.classification_id, "未分类") if f.classification_id else "未分类"
        tier = f.sensitivity_level or "未分级"
        key = f"{cat_name}|{tier}"
        matrix[key] = matrix.get(key, 0) + 1

    return [{"category": k.split("|")[0], "tier": k.split("|")[1], "count": v} for k, v in matrix.items()]


@router.get("/compliance/audit-trail")
def compliance_audit_trail(
    date_from: str | None = None,
    date_to: str | None = None,
    module: str | None = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    q = db.query(OperationLog)
    if date_from:
        q = q.filter(OperationLog.created_at >= date_from)
    if date_to:
        q = q.filter(OperationLog.created_at <= f"{date_to} 23:59:59")
    if module:
        q = q.filter(OperationLog.module == module)

    total = q.count()
    items = q.order_by(OperationLog.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()

    return {
        "items": [
            {"id": i.id, "username": i.username, "action": i.action, "module": i.module,
             "target_type": i.target_type, "target_id": i.target_id, "detail": i.detail,
             "ip_address": i.ip_address, "created_at": i.created_at.isoformat()}
            for i in items
        ],
        "total": total, "page": page, "page_size": page_size,
    }


@router.get("/compliance/tagging-history")
def compliance_tagging_history(
    date_from: str | None = None,
    date_to: str | None = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    q = db.query(TaggingHistory)
    if date_from:
        q = q.filter(TaggingHistory.created_at >= date_from)
    if date_to:
        q = q.filter(TaggingHistory.created_at <= f"{date_to} 23:59:59")

    total = q.count()
    items = q.order_by(TaggingHistory.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()

    field_ids = [i.field_id for i in items]
    field_map = {}
    if field_ids:
        for f in db.query(Field).filter(Field.id.in_(field_ids)).all():
            field_map[f.id] = f.name

    return {
        "items": [
            {"id": i.id, "field_id": i.field_id, "field_name": field_map.get(i.field_id, ""),
             "action": i.action, "old_tier": i.old_tier_level, "new_tier": i.new_tier_level,
             "confidence": i.new_confidence, "method": i.tagging_method, "comment": i.comment,
             "created_at": i.created_at.isoformat()}
            for i in items
        ],
        "total": total, "page": page, "page_size": page_size,
    }


@router.get("/compliance/gaps")
def compliance_gaps(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    unclassified = db.query(Field).filter(
        Field.status == "active", Field.classification_id.is_(None)
    ).count()
    no_tier = db.query(Field).filter(
        Field.status == "active", Field.sensitivity_level.is_(None)
    ).count()
    low_confidence = db.query(Field).filter(
        Field.status == "active", Field.tagging_confidence.isnot(None),
        Field.tagging_confidence < 0.5
    ).count()
    unmapped = db.query(Field).filter(
        Field.status == "active", Field.is_anomaly == True
    ).count()

    gaps = []
    for f in db.query(Field).filter(
        Field.status == "active",
        (Field.classification_id.is_(None)) | (Field.tagging_confidence.isnot(None) & (Field.tagging_confidence < 0.5))
    ).limit(100).all():
        gaps.append({
            "field_id": f.id, "field_code": f.field_code, "name": f.name,
            "table_name": f.table_name,
            "issue": "未分类" if not f.classification_id else "低置信度",
        })

    return {"unclassified": unclassified, "no_tier": no_tier, "low_confidence": low_confidence,
            "unmapped": unmapped, "gaps": gaps}


@router.get("/compliance/export")
def export_compliance_report(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    import io
    import openpyxl

    wb = openpyxl.Workbook()

    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = "合规总览"
    total = db.query(Field).filter(Field.status == "active").count()
    classified = db.query(Field).filter(Field.status == "active", Field.classification_id.isnot(None)).count()
    ws1.append(["指标", "值"])
    ws1.append(["总字段数", total])
    ws1.append(["已分类", classified])
    ws1.append(["覆盖率", f"{round(classified/total*100,1) if total else 0}%"])

    # Sheet 2: Fields detail
    ws2 = wb.create_sheet("字段详情")
    ws2.append(["字段编码", "字段名称", "表名", "分类", "分级", "打标方法", "置信度", "最后打标时间"])
    for f in db.query(Field).filter(Field.status == "active").all():
        cat_name = None
        if f.classification_id:
            cat = db.query(ClassificationCategory).filter(ClassificationCategory.id == f.classification_id).first()
            cat_name = cat.name if cat else None
        ws2.append([f.field_code, f.name, f.table_name, cat_name or "未分类",
                    f.sensitivity_level or "未分级", f.tagging_method or "-",
                    f.tagging_confidence or "-",
                    f.last_tagged_at.isoformat() if f.last_tagged_at else "-"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=compliance_report.xlsx"},
    )
