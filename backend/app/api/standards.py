import json
from datetime import datetime, timezone
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.security import get_current_user, require_role
from app.models.user import User
from app.models.standard import ClassificationCategory, TieringRule
from app.schemas.standard import (
    CategoryCreate, CategoryUpdate, CategoryResponse, CategoryTreeNode,
    TieringRuleCreate, TieringRuleUpdate, TieringRuleResponse,
)
from app.services.log_service import log_action

router = APIRouter(prefix="/api/standards", tags=["Standards"])


# ════════════════════════════════════════════════════════════════════════
# Classification Categories
# ════════════════════════════════════════════════════════════════════════

def _build_tree_node(cat: ClassificationCategory, db: Session) -> CategoryTreeNode:
    children_count = db.query(ClassificationCategory).filter(
        ClassificationCategory.parent_id == cat.id,
        ClassificationCategory.is_active == True,
    ).count()
    return CategoryTreeNode(
        id=cat.id,
        name=cat.name,
        code=cat.code,
        parent_id=cat.parent_id,
        level=cat.level,
        category_type=cat.category_type,
        children_count=children_count,
    )


@router.get("/categories/tree", response_model=list[CategoryTreeNode])
def get_category_tree(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    cats = (
        db.query(ClassificationCategory)
        .filter(ClassificationCategory.is_active == True)
        .order_by(ClassificationCategory.level, ClassificationCategory.sort_order)
        .all()
    )
    return [_build_tree_node(c, db) for c in cats]


@router.get("/categories/", response_model=list[CategoryResponse])
def list_categories(
    category_type: str | None = None,
    version: str | None = None,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    q = db.query(ClassificationCategory).filter(ClassificationCategory.is_active == True)
    if category_type:
        q = q.filter(ClassificationCategory.category_type == category_type)
    if version:
        q = q.filter(ClassificationCategory.version == version)
    return q.order_by(ClassificationCategory.level, ClassificationCategory.sort_order).all()


@router.get("/categories/{cat_id}", response_model=CategoryResponse)
def get_category(cat_id: int, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    cat = db.query(ClassificationCategory).filter(
        ClassificationCategory.id == cat_id, ClassificationCategory.is_active == True
    ).first()
    if not cat:
        raise HTTPException(status_code=404, detail="Category not found")
    return cat


@router.post("/categories/", response_model=CategoryResponse, status_code=201)
def create_category(
    body: CategoryCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("system_admin", "admin")),
    request: Request = None,
):
    if db.query(ClassificationCategory).filter(ClassificationCategory.code == body.code).first():
        raise HTTPException(status_code=400, detail="Code already exists")

    level = 0
    if body.parent_id:
        parent = db.query(ClassificationCategory).filter(
            ClassificationCategory.id == body.parent_id, ClassificationCategory.is_active == True
        ).first()
        if not parent:
            raise HTTPException(status_code=400, detail="Parent category not found")
        level = parent.level + 1

    cat = ClassificationCategory(
        name=body.name, code=body.code, parent_id=body.parent_id, level=level,
        category_type=body.category_type, description=body.description,
        keywords=body.keywords, regulatory_ref=body.regulatory_ref,
        version=body.version, sort_order=body.sort_order, created_by=current_user.id,
    )
    db.add(cat)
    db.commit()
    db.refresh(cat)
    log_action(db, user_id=current_user.id, username=current_user.username,
               action="create", module="standards", target_type="category", target_id=cat.id,
               detail={"name": cat.name, "code": cat.code},
               ip_address=request.client.host if request else None)
    return cat


@router.put("/categories/{cat_id}", response_model=CategoryResponse)
def update_category(
    cat_id: int,
    body: CategoryUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("system_admin", "admin")),
):
    cat = db.query(ClassificationCategory).filter(
        ClassificationCategory.id == cat_id, ClassificationCategory.is_active == True
    ).first()
    if not cat:
        raise HTTPException(status_code=404, detail="Category not found")

    if body.code and body.code != cat.code:
        if db.query(ClassificationCategory).filter(ClassificationCategory.code == body.code).first():
            raise HTTPException(status_code=400, detail="Code already exists")
        cat.code = body.code

    updatable = ["name", "category_type", "description", "keywords", "regulatory_ref", "sort_order"]
    for field in updatable:
        val = getattr(body, field, None)
        if val is not None:
            setattr(cat, field, val)

    if body.parent_id is not None and body.parent_id != cat.parent_id:
        if body.parent_id == cat_id:
            raise HTTPException(status_code=400, detail="Cannot set self as parent")
        new_level = 0
        if body.parent_id:
            parent = db.query(ClassificationCategory).filter(
                ClassificationCategory.id == body.parent_id, ClassificationCategory.is_active == True
            ).first()
            if not parent:
                raise HTTPException(status_code=400, detail="Parent category not found")
            new_level = parent.level + 1
        cat.parent_id = body.parent_id
        cat.level = new_level

    cat.updated_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(cat)
    return cat


@router.delete("/categories/{cat_id}")
def delete_category(
    cat_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(require_role("system_admin", "admin")),
):
    cat = db.query(ClassificationCategory).filter(
        ClassificationCategory.id == cat_id, ClassificationCategory.is_active == True
    ).first()
    if not cat:
        raise HTTPException(status_code=404, detail="Category not found")

    children = db.query(ClassificationCategory).filter(
        ClassificationCategory.parent_id == cat_id, ClassificationCategory.is_active == True
    ).count()
    if children > 0:
        raise HTTPException(status_code=400, detail="Cannot delete category with children")

    cat.is_active = False
    db.commit()
    return {"message": "Category deleted"}


# ── Category Excel Import / Export ──

CATEGORY_HEADERS = ["name", "code", "parent_code", "category_type", "description", "keywords", "regulatory_ref"]
CATEGORY_HEADER_LABELS = ["名称", "编码", "父级编码", "分类类型", "描述", "关键词", "法规参考"]


@router.get("/categories/export")
def export_categories(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    cats = (
        db.query(ClassificationCategory)
        .filter(ClassificationCategory.is_active == True)
        .order_by(ClassificationCategory.level, ClassificationCategory.sort_order)
        .all()
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "分类标准"
    ws.append(CATEGORY_HEADER_LABELS)

    code_to_code = {c.id: c.code for c in cats}
    for c in cats:
        parent_code = code_to_code.get(c.parent_id, "") if c.parent_id else ""
        ws.append([c.name, c.code, parent_code, c.category_type,
                    c.description or "", c.keywords or "", c.regulatory_ref or ""])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=classification_categories.xlsx"})


@router.post("/categories/import")
def import_categories(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("system_admin", "admin")),
):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only .xlsx or .xls files supported")

    from openpyxl import load_workbook
    wb = load_workbook(file.file)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    created = 0
    code_to_id: dict[str, int] = {}
    errors = []

    for i, row in enumerate(rows, start=2):
        if not row or not row[0]:
            continue
        name, code = str(row[0]).strip(), str(row[1]).strip() if row[1] else ""
        parent_code = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        category_type = str(row[3]).strip() if len(row) > 3 and row[3] else "business"
        description = str(row[4]).strip() if len(row) > 4 and row[4] else None
        keywords = str(row[5]).strip() if len(row) > 5 and row[5] else None
        regulatory_ref = str(row[6]).strip() if len(row) > 6 and row[6] else None

        if not name or not code:
            errors.append({"row": i, "error": "名称和编码为必填"})
            continue
        if db.query(ClassificationCategory).filter(ClassificationCategory.code == code).first():
            errors.append({"row": i, "error": f"编码 {code} 已存在"})
            continue

        parent_id = code_to_id.get(parent_code)
        level = 0
        if parent_id:
            parent = db.query(ClassificationCategory).filter(ClassificationCategory.id == parent_id).first()
            if parent:
                level = parent.level + 1

        cat = ClassificationCategory(
            name=name, code=code, parent_id=parent_id, level=level,
            category_type=category_type, description=description,
            keywords=keywords, regulatory_ref=regulatory_ref,
            created_by=current_user.id,
        )
        db.add(cat)
        db.flush()
        code_to_id[code] = cat.id
        created += 1

    db.commit()
    return {"created": created, "errors": errors}


# ════════════════════════════════════════════════════════════════════════
# Tiering Rules
# ════════════════════════════════════════════════════════════════════════

@router.get("/tiers/", response_model=list[TieringRuleResponse])
def list_tiers(
    tier_level: str | None = None,
    rule_type: str | None = None,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    q = db.query(TieringRule).filter(TieringRule.is_active == True)
    if tier_level:
        q = q.filter(TieringRule.tier_level == tier_level)
    if rule_type:
        q = q.filter(TieringRule.rule_type == rule_type)
    return q.order_by(TieringRule.priority.desc()).all()


@router.get("/tiers/{tier_id}", response_model=TieringRuleResponse)
def get_tier(tier_id: int, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    tier = db.query(TieringRule).filter(
        TieringRule.id == tier_id, TieringRule.is_active == True
    ).first()
    if not tier:
        raise HTTPException(status_code=404, detail="Tiering rule not found")
    return tier


@router.post("/tiers/", response_model=TieringRuleResponse, status_code=201)
def create_tier(
    body: TieringRuleCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("system_admin", "admin")),
):
    tier = TieringRule(
        tier_level=body.tier_level, tier_name=body.tier_name,
        rule_type=body.rule_type, rule_content=body.rule_content,
        priority=body.priority, regulatory_basis=body.regulatory_basis,
        version=body.version, created_by=current_user.id,
    )
    db.add(tier)
    db.commit()
    db.refresh(tier)
    return tier


@router.put("/tiers/{tier_id}", response_model=TieringRuleResponse)
def update_tier(
    tier_id: int,
    body: TieringRuleUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("system_admin", "admin")),
):
    tier = db.query(TieringRule).filter(
        TieringRule.id == tier_id, TieringRule.is_active == True
    ).first()
    if not tier:
        raise HTTPException(status_code=404, detail="Tiering rule not found")

    updatable = ["tier_level", "tier_name", "rule_type", "rule_content", "priority", "regulatory_basis", "is_active"]
    for field in updatable:
        val = getattr(body, field, None)
        if val is not None:
            setattr(tier, field, val)

    tier.updated_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(tier)
    return tier


@router.delete("/tiers/{tier_id}")
def delete_tier(
    tier_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(require_role("system_admin", "admin")),
):
    tier = db.query(TieringRule).filter(
        TieringRule.id == tier_id, TieringRule.is_active == True
    ).first()
    if not tier:
        raise HTTPException(status_code=404, detail="Tiering rule not found")
    tier.is_active = False
    db.commit()
    return {"message": "Tiering rule deleted"}


# ── Tier Excel Import / Export ──

TIER_HEADERS = ["tier_level", "tier_name", "rule_type", "rule_content", "priority", "regulatory_basis"]
TIER_HEADER_LABELS = ["分级级别", "分级名称", "规则类型", "规则内容(JSON)", "优先级", "法规依据"]


@router.get("/tiers/export")
def export_tiers(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    tiers = (
        db.query(TieringRule)
        .filter(TieringRule.is_active == True)
        .order_by(TieringRule.priority.desc())
        .all()
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "分级规则"
    ws.append(TIER_HEADER_LABELS)

    for t in tiers:
        ws.append([t.tier_level, t.tier_name, t.rule_type,
                    t.rule_content, t.priority, t.regulatory_basis or ""])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=tiering_rules.xlsx"})


@router.post("/tiers/import")
def import_tiers(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("system_admin", "admin")),
):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only .xlsx or .xls files supported")

    from openpyxl import load_workbook
    wb = load_workbook(file.file)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    created = 0
    errors = []

    for i, row in enumerate(rows, start=2):
        if not row or not row[0]:
            continue
        tier_level = str(row[0]).strip() if row[0] else ""
        tier_name = str(row[1]).strip() if row[1] else ""
        rule_type = str(row[2]).strip() if len(row) > 2 and row[2] else "keyword"
        rule_content = str(row[3]).strip() if len(row) > 3 and row[3] else "{}"
        priority = int(row[4]) if len(row) > 4 and row[4] else 0
        regulatory_basis = str(row[5]).strip() if len(row) > 5 and row[5] else None

        if not tier_level or not tier_name:
            errors.append({"row": i, "error": "分级级别和名称为必填"})
            continue

        tier = TieringRule(
            tier_level=tier_level, tier_name=tier_name,
            rule_type=rule_type, rule_content=rule_content,
            priority=priority, regulatory_basis=regulatory_basis,
            created_by=current_user.id,
        )
        db.add(tier)
        created += 1

    db.commit()
    return {"created": created, "errors": errors}
